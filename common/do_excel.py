from openpyxl import load_workbook

test_data = []


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def do_excel(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        for i in range(2, sheet.max_row+1):
            sub_data = {}
            sub_data['id'] = sheet.cell(i, 1).value
            sub_data['title'] = sheet.cell(i, 2).value
            sub_data['method'] = sheet.cell(i, 3).value
            sub_data['url'] = sheet.cell(i, 4).value
            sub_data['params'] = sheet.cell(i, 5).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    test_data = DoExcel(r'../test_data/data.xlsx', 'Sheet1').do_excel()

    for item in test_data:
        print(item)
